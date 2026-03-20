from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime
from openpyxl import Workbook, load_workbook
from fastapi.responses import StreamingResponse
import io

from app.database import get_db
from app.auth_utils import get_current_user
from app.models import (
    TonKhoChotNgay,
    QuyNhanVienChotNgay,
    QuyCongTyChotNgay,
    ThuChi
)
from app.schemas import KhoiTaoDauKyRequest

router = APIRouter(prefix="/system", tags=["system"])


# ====================================================
# 1. KHỞI TẠO ĐẦU KỲ
# ====================================================

@router.post("/khoi-tao-dau-ky")
def khoi_tao_dau_ky(
    data: KhoiTaoDauKyRequest,
    db: Session = Depends(get_db),
    user = Depends(get_current_user)
):
    if user.vai_tro != "admin":
        raise HTTPException(403, "Chỉ admin")

    try:
        with db.begin():

            # ❌ không cho chạy lại
            if db.query(QuyCongTyChotNgay).first():
                raise HTTPException(400, "Đã khởi tạo rồi")

            # ===== TỒN KHO =====
            for item in data.ton_kho:
                db.add(TonKhoChotNgay(
                    ma_sp=item.ma_sp,
                    ma_kho=item.ma_kho,
                    so_luong=item.so_luong
                ))

            # ===== QUỸ NV =====
            for q in data.quy_nhan_vien:
                db.add(QuyNhanVienChotNgay(
                    ma_nv=q.ma_nv,
                    so_du=q.so_du
                ))

            # ===== QUỸ CTY =====
            db.add(QuyCongTyChotNgay(
                tien_mat=data.quy_cong_ty,
                tien_ngan_hang=0,
                tong_quy=data.quy_cong_ty
            ))

        return {"message": "Khởi tạo OK"}

    except Exception as e:
        raise HTTPException(500, str(e))


# ====================================================
# 2. EXPORT EXCEL
# ====================================================

@router.get("/export-dau-ky")
def export_dau_ky(
    db: Session = Depends(get_db),
    user = Depends(get_current_user)
):
    if user.vai_tro != "admin":
        raise HTTPException(403, "Chỉ admin")

    wb = Workbook()

    # ===== TỒN KHO =====
    ws = wb.active
    ws.title = "ton_kho"
    ws.append(["ma_kho", "ma_sp", "so_luong"])

    for row in db.query(TonKhoChotNgay).all():
        ws.append([row.ma_kho, row.ma_sp, float(row.so_luong)])

    # ===== QUỸ NV =====
    ws = wb.create_sheet("quy_nhan_vien")
    ws.append(["ma_nv", "so_du"])

    for q in db.query(QuyNhanVienChotNgay).all():
        ws.append([q.ma_nv, float(q.so_du)])

    # ===== QUỸ CTY =====
    ws = wb.create_sheet("quy_cong_ty")
    ws.append(["tien_mat", "tien_ngan_hang"])

    ct = db.query(QuyCongTyChotNgay).first()
    if ct:
        ws.append([float(ct.tien_mat), float(ct.tien_ngan_hang)])

    # ===== STREAM FILE =====
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dau_ky.xlsx"}
    )


# ====================================================
# 3. IMPORT EXCEL (RESET DB)
# ====================================================

@router.post("/import-dau-ky")
def import_dau_ky(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user = Depends(get_current_user)
):
    if user.vai_tro != "admin":
        raise HTTPException(403, "Chỉ admin")

    try:
        wb = load_workbook(file.file)

        with db.begin():

            # ❌ chặn nếu đã phát sinh giao dịch
            if db.query(ThuChi).count() > 0:
                raise HTTPException(400, "Đã có giao dịch, không cho import")

            # ===== RESET =====
            db.query(TonKhoChotNgay).delete()
            db.query(QuyNhanVienChotNgay).delete()
            db.query(QuyCongTyChotNgay).delete()

            # ===== TỒN KHO =====
            ws = wb["ton_kho"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                ma_kho, ma_sp, so_luong = row

                if not ma_sp:
                    continue

                db.add(TonKhoChotNgay(
                    ma_kho=ma_kho,
                    ma_sp=ma_sp,
                    so_luong=so_luong or 0
                ))

            # ===== QUỸ NV =====
            ws = wb["quy_nhan_vien"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                ma_nv, so_du = row

                if not ma_nv:
                    continue

                db.add(QuyNhanVienChotNgay(
                    ma_nv=ma_nv,
                    so_du=so_du or 0
                ))

            # ===== QUỸ CTY =====
            ws = wb["quy_cong_ty"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                tien_mat, tien_ngan_hang = row

                db.add(QuyCongTyChotNgay(
                    tien_mat=tien_mat or 0,
                    tien_ngan_hang=tien_ngan_hang or 0,
                    tong_quy=(tien_mat or 0) + (tien_ngan_hang or 0)
                ))

        return {"message": "Import thành công"}

    except Exception as e:
        raise HTTPException(400, str(e))
